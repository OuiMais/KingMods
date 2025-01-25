"""
    Projet : KingMods
    Date Creation : 07/08/2023
    Date Revision : 13/01/2025
    Entreprise : 3SC4P3
    Auteur: Florian HOFBAUER
    Contact :
    But : Obtenir un document comportant tous les liens et noms des derniers mods ajoutés sur KingMods dans le but de générer un bot
"""
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import os


def ModResumeWithLastModSave():
    """
        Function to save all title and link of KingsMod's mods until we find the last mod we see.
        To start, recuperation of last title mod from New and Update mod.
        Next, recuperation of mod from KingsMod.
        Last, save an excel file.
    """

    if os.path.exists("KingsMod_Resume.xlsx"):
        excelPrevious = openpyxl.open("KingsMod_Resume.xlsx")
        excelSheetPrevious = excelPrevious.active
        previousNewTitle = excelSheetPrevious["A2"].value
        sheetNames = excelPrevious.sheetnames
        excelSheetPrevious = excelPrevious[sheetNames[1]]
        previousUpdateTitle = excelSheetPrevious["A2"].value
        excelPrevious.close()
    else:
        previousUpdateTitle = None
        previousNewTitle = None

    if previousUpdateTitle is None:
        previousUpdateTitle = "Hangar à machines 55x74"
    if previousNewTitle is None:
        previousNewTitle = "Hangar à machines 55x74"

    url = "https://www.kingmods.net/fr/fs25/nouveaux-mods"

    modTitleAndLink = []
    pageModArray = []

    findPrevious = 0
    page = 1

    # Initiate the browser
    options = Options()
    options.add_argument('--headless')
    browser = webdriver.Chrome(options=options)

    # Open website
    browser.get(url)

    while not findPrevious:
        # Save ul data
        pageModList = browser.find_element(By.XPATH, "//ul[@class='w-full grid gap-20 grid-cols-2']")

        lastPageMod = pageModList.find_elements(By.TAG_NAME, 'a')

        for mod in lastPageMod:
            try:
                # Check if badge update is display
                badge = mod.find_element(By.XPATH, ".//div[contains(@class, 'bg-blue')]")
                updateMod = badge.is_displayed()
            except NoSuchElementException:
                updateMod = False

            title = mod.get_attribute("title")
            link = mod.get_attribute("href")

            if title != previousNewTitle and title != previousUpdateTitle:
                pageModArray.append([title, link, updateMod])
            else:
                modTitleAndLink += pageModArray
                findPrevious = 1
                break

        page += 1
        url = "https://www.kingmods.net/fr/fs25/nouveaux-mods?page=" + str(page)

        browser.get(url)

    browser.close()

    filteredMod = []
    seen = set()

    for mod in modTitleAndLink:
        key = (mod[0], mod[1])
        if key not in seen:
            seen.add(key)
            filteredMod.append(mod)

    if len(filteredMod) == 0:
        filteredMod.append([previousNewTitle, "", 0])
        filteredMod.append([previousUpdateTitle, "", 1])

    SaveExcelFile("KingsMod_Resume.xlsx", filteredMod)


def LastDayModResume():
    """
        Function to resume all mod add or modify on KingsMod
    """
    url = "https://www.kingmods.net/fr/fs25/nouveaux-mods"

    modTitleAndLink = []
    pageModArray = []

    findLast = 0
    page = 1

    today = datetime.date.today()
    dayToStop = today - datetime.timedelta(days=2)
    dayToStop = dayToStop.strftime("%Y-%m-%d")

    # Initiate the browser
    options = Options()
    options.add_argument('--headless')
    browser = webdriver.Chrome(options=options)

    # Open website
    browser.get(url)

    while not findLast:
        # Save ul data
        pageModList = browser.find_element(By.XPATH, "//ul[@class='w-full grid gap-20 grid-cols-2']")

        lastPageMod = pageModList.find_elements(By.TAG_NAME, 'a')

        for mod in lastPageMod:
            try:
                # Check if badge update is display
                badge = mod.find_element(By.XPATH, ".//div[contains(@class, 'bg-blue')]")
                updateMod = badge.is_displayed()
            except NoSuchElementException:
                updateMod = False

            timeMod = mod.find_element(By.TAG_NAME, "time")
            dateTimeMod = timeMod.get_attribute("datetime")
            modDay = dateTimeMod.split("T")[0]

            title = mod.get_attribute("title")
            link = mod.get_attribute("href")

            if modDay != dayToStop:
                pageModArray.append([title, link, updateMod])
            else:
                modTitleAndLink += pageModArray
                findLast = 1
                break

        page += 1
        url = "https://www.kingmods.net/fr/fs25/nouveaux-mods?page=" + str(page)

        browser.get(url)

    browser.close()

    filteredMod = []
    seen = set()

    for mod in modTitleAndLink:
        key = (mod[0], mod[1])
        if key not in seen:
            seen.add(key)
            filteredMod.append(mod)

    filepath = "KingMod_Resume_" + today.strftime("%Y-%m-%d") + ".xlsx"
    SaveExcelFile(filepath, filteredMod)


def SaveExcelFile(filepath, modList):
    """
        Function to save mod in excel file

        Args:
            - filepath : [str] : filepath to xcl file
            - modList : table with title and link of mod: [title, link]
    """
    excelFile = Workbook()

    newModSheet = excelFile.active
    newModSheet.title = "New Mod"
    updateModSheet = excelFile.create_sheet(title="Update Mod")

    newModSheet.append(["Title", "Link"])
    updateModSheet.append(["Title", "Link"])

    newModSheet.column_dimensions['A'].width = 71
    newModSheet.column_dimensions['B'].width = 142

    updateModSheet.column_dimensions['A'].width = 71
    updateModSheet.column_dimensions['B'].width = 142

    cells = [newModSheet["A1"], newModSheet["B1"], updateModSheet["A1"], updateModSheet["B1"], ]

    for cell in cells:
        cell.font = Font(name="Calibri", size=18, bold=True, italic=False)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    newIndex = 2
    updateIndex = 2

    for mod in modList:
        title = mod[0]
        link = mod[1]

        if mod[2]:
            updateModSheet[f"A{updateIndex}"] = title
            updateModSheet[f"B{updateIndex}"] = f'=HYPERLINK("{link}", "{link}") \n'
            updateIndex += 1
        else:
            newModSheet[f"A{newIndex}"] = title
            newModSheet[f"B{newIndex}"] = f'=HYPERLINK("{link}", "{link}") \n'
            newIndex +=1

    excelFile.save(filepath)


ModResumeWithLastModSave()
