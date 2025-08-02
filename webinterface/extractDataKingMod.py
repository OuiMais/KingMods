"""
    Projet : KingMods
    Date Creation : 07/08/2023
    Date Revision : 02/08/2025
    Entreprise : 3SC4P3
    Auteur: Florian HOFBAUER
    Contact :
    But : Extrait tous les mods de la veille publiés sur le site kingmods
"""
import functools
import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import os
import config, configJson


# Retry decorator
def retry_on_exception(max_retries=3, delay=5, exceptions=(Exception,), task_name=""):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(1, max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except exceptions as e:
                    print(f"[{task_name}] ERREUR (tentative {attempt}/{max_retries}): {e}")
                    if attempt < max_retries:
                        time.sleep(delay)
                    else:
                        print(f"[{task_name}] ÉCHEC APRÈS {max_retries} TENTATIVES.")
        return wrapper
    return decorator

@retry_on_exception(task_name="LANCEMENT BROWSER")
def launchBrowser():
    options = Options()
    options.add_argument('--headless')
    options.add_argument("-disable-gpu")
    options.add_argument("--disable-popup-blocking")

    path = '/usr/bin/chromedriver'
    service = Service(executable_path=path)
    browser = webdriver.Chrome(service=service, options=options)
    return browser

@retry_on_exception(task_name="CHARGEMENT PAGE")
def loadPage(browser, page_url):
    browser.get(page_url)
    return browser.find_element(By.XPATH, "//ul[@class='w-full grid gap-20 grid-cols-2']").find_elements(By.TAG_NAME, 'a')

@retry_on_exception(task_name="ENVOI EMAIL")
def sendMail(filepath, sender_email, recipient_email, password, message):
    with open(filepath, "rb") as file:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={filepath.split('/')[-1]}")
        message.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender_email, password)
        server.sendmail(sender_email, recipient_email, message.as_string())

# ####################################################
#                      Variables
# ####################################################
url = "https://www.kingmods.net/fr/fs25/nouveaux-mods"

modTitleAndLink = []
pageModArray = []

findLast = 0
page = 1

today = datetime.date.today()
dayToStopDate = today - datetime.timedelta(days=2)
dayToStop = dayToStopDate.strftime("%Y-%m-%d")

# ####################################################
#                 Initiate the browser
# ####################################################

# Initiate the browser
browser = launchBrowser()

# Open website
browser.get(url)

# ####################################################
#                 Exctract mods
# ####################################################
while not findLast:
    # Save ul data
    lastPageMod = loadPage(browser, url)

    for mod in lastPageMod:
        try:
            # Check if badge update is display
            badge = mod.find_element(By.XPATH, ".//div[contains(@class, 'bg-blue')]")
            updateMod = badge.is_displayed()
        except NoSuchElementException:
            updateMod = False

        timeMod = mod.find_element(By.TAG_NAME, "time")
        dateTimeMod = timeMod.get_attribute("datetime")
        modDay = dateTimeMod.split("T")[0]
        modDayDate = datetime.datetime.strptime(modDay, "%Y-%m-%d").date()

        title = mod.get_attribute("title")
        link = mod.get_attribute("href")

        if modDayDate > dayToStopDate:
            pageModArray.append([title, link, updateMod])
        else:
            modTitleAndLink += pageModArray
            findLast = 1
            break

    page += 1
    url = "https://www.kingmods.net/fr/fs25/nouveaux-mods?page=" + str(page)

    browser.get(url)

browser.close()

# ####################################################
#   Remove mods if it's a new mod and an update mod
# ####################################################
filteredMod = []
seen = set()

for mod in modTitleAndLink:
    key = (mod[0], mod[1])
    if key not in seen:
        seen.add(key)
        filteredMod.append(mod)

# ####################################################
#                 Save in xlsx
# ####################################################
# Prepare file
dateOfFile = today - datetime.timedelta(days=1)
filepath = "KingMod_Resume_" + dateOfFile.strftime("%Y_%m_%d") + ".xlsx"

excelFile = Workbook()

newModSheet = excelFile.active
newModSheet.title = "New Mod"
updateModSheet = excelFile.create_sheet(title="Update Mod")

newModSheet.append(["Title", "Link"])
updateModSheet.append(["Title", "Link"])

newModSheet.column_dimensions['A'].width = 71
newModSheet.column_dimensions['B'].width = 142

updateModSheet.column_dimensions['A'].width = 71
updateModSheet.column_dimensions['B'].width = 142

cells = [newModSheet["A1"], newModSheet["B1"], updateModSheet["A1"], updateModSheet["B1"]]

for cell in cells:
    cell.font = Font(name="Calibri", size=18, bold=True, italic=False)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Inject data mods
newIndex = 2
updateIndex = 2
youHaveToUpdate = 0

downloadModsData = configJson.load_data(configJson.DOWNLOADED_MODS_JSON_FILE)

newModsData = []

for mod in filteredMod:
    title = mod[0]
    link = mod[1]

    if mod[2]:
        updateModSheet[f"A{updateIndex}"] = title
        updateModSheet[f"B{updateIndex}"] = f'=HYPERLINK("{link}", "{link}") \n'

        for modDl in downloadModsData:
            if modDl["link"] == link:
                updateModSheet[f"A{updateIndex}"].fill = PatternFill(start_color='FFD970', end_color='FFD970',
                                                                     fill_type="solid")
                updateModSheet[f"B{updateIndex}"].fill = PatternFill(start_color='FFD970', end_color='FFD970',
                                                                     fill_type="solid")
                youHaveToUpdate += 1
                modDl["toUpdate"] = 1
                break
        updateIndex += 1

    else:
        newModSheet[f"A{newIndex}"] = title
        newModSheet[f"B{newIndex}"] = f'=HYPERLINK("{link}", "{link}") \n'
        newIndex +=1

        newModsData.append({"title": title, "link": link})

configJson.save_data(configJson.DOWNLOADED_MODS_JSON_FILE, downloadModsData)
configJson.save_data(configJson.NEW_MODS_JSON_FILE, newModsData)

excelFile.save(filepath)
excelFile.close()

# ####################################################
#                      Send mail
# ####################################################
sender_email = config.sender_email
recipient_email = config.recipient_email
password = config.password

# Configuration de l'email
message = MIMEMultipart()
message["From"] = sender_email
message["To"] = recipient_email
message["Subject"] = "KingMod_Day_" + dateOfFile.strftime("%Y_%m_%d")

text = ("Bonjour,\n\nAujourd'hui, il y a eu " + str(newIndex - 2) + " nouveaux mods et " + str(updateIndex - 2) +
        " mods mis à jour.\n\nVous avez " + str(youHaveToUpdate) + " mods to update.")
message.attach(MIMEText(text, "plain"))

sendMail(filepath, sender_email, recipient_email, password, message)

try:
    if os.path.exists(filepath):
        os.remove(filepath)
except Exception as e:
    print(f"[ERREUR SUPPRESSION FICHIER] : {e}")
